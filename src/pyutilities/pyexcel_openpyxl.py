# !/usr/bin/python3
# -*- coding: UTF-8 -*-
""" pyexcel_openpyxl
"""
from typing import override
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet as oWorkSheet

from src.pyutilities.logit import pv
from src.pyutilities.pyexcel import PySheet, PyExcel


class WorkSheet(PySheet):
    def __init__(self, sht: oWorkSheet):
        self._sheet: oWorkSheet = sht

    @override
    def set_cell(self, cell: str | tuple[int, int], val: int | float | str):
        cell = self._tuple_to_str(cell)
        self._sheet[cell] = val

    @override
    def get_cell(self, cell: str | tuple[int, int]) -> int | float | str | None:
        cell = self._tuple_to_str(cell)
        return self._sheet[cell]

    def _tuple_to_str(self, cell: str | tuple[int, int]) -> str:
        if isinstance(cell, tuple):
            x = cell[0]
            y = cell[1]
            cell = f"{chr(ord('A') + y - 1)}{x}"
        else:
            cell = cell.upper()
        # pv(cell)
        return cell


class Excel(PyExcel):
    def __init__(self, excelfile: str):
        super().__init__(excelfile)
        self._workbook: openpyxl.Workbook = openpyxl.load_workbook(excelfile)

    @override
    def save(self, newfile: str = ""):
        if newfile:
            pv(newfile)
            self._workbook.save(newfile)
        else:
            pv(self._excelfile)
            self._workbook.save(self._excelfile)

    # @override
    # def close(self, save: bool | str = True):
        # if isinstance(save, str):
            # pv(save)
            # self.save(save)
        # elif save:
            # self.save()
        # self._workbook.close()

    @override
    def close(self, save: bool = True):
        if save:
            self.save()
        self._workbook.close()

    @override
    def copy_sheet(self, src_sheet: str, dest_sheet: str) -> WorkSheet:
        copied_sheet = self._workbook.copy_worksheet(self._workbook[src_sheet])
        copied_sheet.title = dest_sheet
        return WorkSheet(copied_sheet)

    @override
    def sheets(self) -> list[str]:
        sheet_list: list[str] = []
        for sht in self._workbook.worksheets:
            sheet_list.append(sht.title)
        return sheet_list

    @override
    def add_sheet(self, sheetname: str = "", location: int | None = None):
        if sheetname:
            ws = self._workbook.create_sheet(sheetname, location)
        else:
            ws = self._workbook.create_sheet()
        return WorkSheet(ws)

    @override
    def get_sheet(self, sheetname: str = ""):
        # get_sheet_by_name
        if sheetname:
            ws = self._workbook[sheetname]
        else:
            ws = self._workbook.active
        return WorkSheet(ws)

    @override
    def remove_sheet(self, sheetname: str = "") -> bool:
        del self._workbook[sheetname]
        return True

    @override
    def rename_sheet(self, old_sheetname: str, new_sheetname: str) -> bool:
        ws = self._workbook[old_sheetname]
        ws.title = new_sheetname
        return True
