# !/usr/bin/python3
# -*- coding: UTF-8 -*-
""" pyexcel_openpyxl
"""
from typing import override
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet as oWorkSheet

try:
    from logit import pv
    from pyexcel import PySheet, PyExcel
except ImportError:
    from pyutilies.logit import pv
    from pyutilies.pyexcel import PySheet, PyExcel


class WorkSheet(PySheet):
    def __init__(self, sht: oWorkSheet):
        self._sheet: oWorkSheet = sht

    @override
    def set_cell(self, cell: str | tuple[int, int], val: int | float | str):
        cell = self._tuple_to_str(cell)
        self._sheet[cell] = val

    @override
    def get_cell(self, cell: str | tuple[int, int]) -> int | float | str | None:
        cell = self._tuple_to_str(cell)
        return self._sheet[cell]

    def _tuple_to_str(self, cell: str | tuple[int, int]) -> str:
        if isinstance(cell, tuple):
            x = cell[0]
            y = cell[1]
            cell = f"{chr(ord('A') + y - 1)}{x}"
        else:
            cell = cell.upper()
        # pv(cell)
        return cell


class Excel(PyExcel):
    def __init__(self, excelfile: str):
        super().__init__(excelfile)
        self._workbook: openpyxl.Workbook = openpyxl.load_workbook(excelfile)

    @override
    def save(self, newfile: str = ""):
        if newfile:
            pv(newfile)
            self._workbook.save(newfile)
        else:
            pv(self._excelfile)
            self._workbook.save(self._excelfile)

    # @override
    # def close(self, save: bool | str = True):
        # if isinstance(save, str):
            # pv(save)
            # self.save(save)
        # elif save:
            # self.save()
        # self._workbook.close()

    @override
    def close(self, save: bool = True):
        if save:
            self.save()
        self._workbook.close()

    @override
    def copy_sheet(self, src_sheet: str, dest_sheet: str) -> WorkSheet:
        copied_sheet = self._workbook.copy_worksheet(self._workbook[src_sheet])
        copied_sheet.title = dest_sheet
        return WorkSheet(copied_sheet)

    @override
    def sheets(self) -> list[str]:
        sheet_list: list[str] = []
        for sht in self._workbook.worksheets:
            sheet_list.append(sht.title)
        return sheet_list

    @override
    def add_sheet(self, sheetname: str = "", location: int | None = None):
        if sheetname:
            ws = self._workbook.create_sheet(sheetname, location)
        else:
            ws = self._workbook.create_sheet()
        return WorkSheet(ws)

    @override
    def get_sheet(self, sheetname: str = ""):
        # get_sheet_by_name
        if sheetname:
            ws = self._workbook[sheetname]
        else:
            ws = self._workbook.active
        return WorkSheet(ws)

    @override
    def remove_sheet(self, sheetname: str = "") -> bool:
        del self._workbook[sheetname]
        return True

    @override
    def rename_sheet(self, old_sheetname: str, new_sheetname: str) -> bool:
        ws = self._workbook[old_sheetname]
        ws.title = new_sheetname
        return True


if __name__ == '__main__':
    import os
    import sys
    import shutil
    path = os.path.dirname(os.path.abspath(__file__))
    if getattr(sys, "frozen", False):
        path = os.path.dirname(os.path.abspath(sys.executable))

    source_file = os.path.join(path, "resources", "template.xlsx")
    dest_file = os.path.join(path, "resources", "logs", "123.xlsx")

    try:
        _ = shutil.copyfile(source_file, dest_file)
    except OSError as e:
        print(f"Unable to copy file, {e}")

    excel = Excel(dest_file)
    new_sheet = "ISG_26_01"
    ws = excel.copy_sheet("ISG", new_sheet)

    start_index_adr = "E3"
    single_gap_adr = 2, 2       # B2
    line_gap_adr = 2, 1         # A2
    frame_adr = 3, 1            # A3

    data_isg = os.path.join(path, "resources", "data.txt")
    with open(data_isg, encoding='utf-8') as f:
        data = f.read()

    ws.set_cell(start_index_adr, 28)
    ws.set_cell(single_gap_adr, 3)
    ws.set_cell(line_gap_adr, 190)
    ws.set_cell(frame_adr, data)

    excel.close()
